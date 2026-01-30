# src/ui/enseignant_window.py
"""
Fenêtre principale de l'enseignant
Fonctionnalités : Emploi du temps, Réservations, Recherche Salles, Indisponibilités
"""

import os
import csv
from datetime import datetime

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
    QPushButton, QFrame, QStackedWidget, QTableWidget, 
    QTableWidgetItem, QHeaderView, QComboBox, QDateEdit,
    QTimeEdit, QTextEdit, QMessageBox, QGridLayout
)
from PyQt6.QtCore import Qt, QDate, QTime, pyqtSignal
from PyQt6.QtGui import QIcon, QPixmap, QColor

from configUI import WINDOW_CONFIG, COLORS, FST_LOGO_IMAGE
from config import MATIERES_COMPLETES, SPECIALITE_KEYWORDS
from src.ui.styles import (
    GLOBAL_STYLE, SIDEBAR_STYLE, SIDEBAR_BUTTON_STYLE, 
    SIDEBAR_USER_INFO_STYLE, CARD_STYLE, CARD_TITLE_STYLE,
    PRIMARY_BUTTON_STYLE, SECONDARY_BUTTON_STYLE,
    TABLE_STYLE, INPUT_STYLE, DANGER_BUTTON_STYLE
)

class UserWrapper:
    def __init__(self, user_tuple):
        self.id = user_tuple[0]
        self.nom = user_tuple[1]
        self.prenom = user_tuple[2]
        self.email = user_tuple[3]
        self.specialite = user_tuple[6] if len(user_tuple) > 6 else None
        # Map other fields if necessary

class EnseignantWindow(QWidget):
    logout_signal = pyqtSignal()

    def __init__(self, user, db):
        super().__init__()
        # Handle tuple vs object
        if isinstance(user, tuple):
            self.user = UserWrapper(user)
        else:
            self.user = user
            
        self.db = db
        
        self.setWindowTitle(WINDOW_CONFIG['enseignant']['title'])
        self.setMinimumSize(1024, 768)
        self.showMaximized()
        self.setStyleSheet(GLOBAL_STYLE)
        
        # Layout principal
        self.main_layout = QHBoxLayout(self)
        self.main_layout.setContentsMargins(0, 0, 0, 0)
        self.main_layout.setSpacing(0)
        
        self.create_sidebar()
        self.create_content_area()
        self.switch_page("Emploi du Temps")

    def create_sidebar(self):
        """Menu latéral Enseignant"""
        self.sidebar = QFrame()
        self.sidebar.setFixedWidth(280)
        self.sidebar.setStyleSheet(SIDEBAR_STYLE)
        
        layout = QVBoxLayout(self.sidebar)
        layout.setContentsMargins(0, 0, 0, 20)
        
        # Header
        header = QFrame()
        header.setStyleSheet(f"background-color: {COLORS['primary_blue']};")
        h_layout = QVBoxLayout(header)
        h_layout.setContentsMargins(20, 30, 20, 30)
        

            
        title = QLabel("ESPACE ENSEIGNANT")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("color: white; font-weight: bold; margin-top: 10px;")
        h_layout.addWidget(title)
        
        layout.addWidget(header)
        
        # Menu
        self.menu_buttons = {}
        menus = [
            ("Emploi du Temps", "Schedule"),
            ("Réserver une séance", "Reservation"), # Renamed from SearchSession
            ("Indisponibilités", "Unavailability")
        ]
        
        for label, pid in menus:
            btn = QPushButton(f"  {label}")
            btn.setCheckable(True)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setStyleSheet(SIDEBAR_BUTTON_STYLE)
            btn.clicked.connect(lambda _, p=pid: self.switch_page(p))
            layout.addWidget(btn)
            self.menu_buttons[pid] = btn
            
        layout.addStretch()
        
        # Footer
        user_lbl = QLabel(f"Prof. {self.user.nom if self.user else 'Enseignant'}")
        user_lbl.setStyleSheet(SIDEBAR_USER_INFO_STYLE)
        user_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(user_lbl)
        
        logout = QPushButton("Déconnexion")
        logout.setCursor(Qt.CursorShape.PointingHandCursor)
        logout.setStyleSheet("""
            QPushButton {
                background: transparent; color: #e74c3c; border: 1px solid #e74c3c;
                border-radius: 5px; padding: 8px; margin: 15px; font-weight: bold;
            }
            QPushButton:hover { background: #e74c3c; color: white; }
        """)
        logout.clicked.connect(self.logout_signal.emit)
        layout.addWidget(logout)
        
        self.main_layout.addWidget(self.sidebar)

    def create_content_area(self):
        self.content = QFrame()
        self.content.setStyleSheet(f"background-color: {COLORS['bg_light']};")
        layout = QVBoxLayout(self.content)
        layout.setContentsMargins(40, 40, 40, 40)
        
        self.page_title = QLabel("Mon Emploi du Temps")
        self.page_title.setStyleSheet(f"font-size: 28px; font-weight: bold; color: {COLORS['text_dark']}; margin-bottom: 20px;")
        layout.addWidget(self.page_title)
        
        self.pages = QStackedWidget()
        self.pages.addWidget(self.create_schedule_page())
        self.pages.addWidget(self.create_reservation_page()) 
        self.pages.addWidget(self.create_unavailability_page())
        # Removed create_search_page (Room Search)
        
        layout.addWidget(self.pages)
        self.main_layout.addWidget(self.content)

    def create_schedule_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        
        # Info enseignant
        info_frame = QFrame()
        info_frame.setStyleSheet("""
            background-color: white; 
            border-radius: 10px; 
            padding: 15px;
            border: 1px solid #E0E0E0;
        """)
        info_layout = QHBoxLayout(info_frame)
        
        specialite = getattr(self.user, 'specialite', None) or "Non définie"
        info_str = f"""
        <div style='font-size: 14px; color: {COLORS['text_dark']};'>
            <b>Enseignant:</b> {self.user.prenom} {self.user.nom}<br/>
            <b>Email:</b> {self.user.email}<br/>
            <b>Spécialité:</b> {specialite}
        </div>
        """
        info_label = QLabel(info_str)
        info_label.setTextFormat(Qt.TextFormat.RichText)
        info_layout.addWidget(info_label)
        info_layout.addStretch()
        
        # Actions (Exporter)
        for fmt in ["PDF", "Excel", "Image"]:
            btn = QPushButton(f"Exporter {fmt}")
            btn.setStyleSheet(SECONDARY_BUTTON_STYLE)
            btn.clicked.connect(lambda _, f=fmt: self.export_schedule(f))
            info_layout.addWidget(btn)
            
        layout.addWidget(info_frame)
        
        # Table
        self.schedule_table = QTableWidget(5, 6)
        self.schedule_table.setHorizontalHeaderLabels(["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi"])
        time_slots = ["08:30 - 10:00", "10:15 - 11:45", "12:00 - 13:30", "13:45 - 15:15", "15:30 - 17:00"]
        self.schedule_table.setVerticalHeaderLabels(time_slots)
        self.schedule_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.schedule_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.schedule_table.setStyleSheet(TABLE_STYLE)
        
        # Load Real Data
        self.load_schedule()
        
        layout.addWidget(self.schedule_table)
        return page

    def load_schedule(self):
        """Charge l'emploi du temps de l'enseignant avec détails complets"""
        self.schedule_table.clearContents()
        
        def get_row(time_str):
            if "08" in time_str or "09" in time_str: return 0
            if "10" in time_str or "11" in time_str: return 1
            if "12" in time_str or "13" in time_str: return 2
            if "14" in time_str or "15" in time_str: return 3
            if "16" in time_str or "17" in time_str: return 4
            return -1

        try:
            seances = self.db.get_seances_by_enseignant(self.user.id)
            
            for s in seances:
                # s: id(0), titre(1), type_seance(2), date(3), h_debut(4), h_fin(5), salle_id(6), enseignant_id(7), groupe_id(8)
                date_str = s[3]
                qdate = QDate.fromString(date_str, "yyyy-MM-dd")
                day_idx = qdate.dayOfWeek() - 1
                
                if 0 <= day_idx <= 5:
                    row = get_row(s[4])
                    if row != -1:
                        # Récupérer le nom de la salle
                        salle_name = "Salle ?"
                        if s[6]:
                            salle = self.db.get_salle_by_id(s[6])
                            if salle:
                                salle_name = salle[1]
                        
                        # Récupérer le nom du groupe
                        groupe_name = ""
                        if s[8]:
                            groupe = self.db.get_groupe_by_id(s[8])
                            if groupe:
                                groupe_name = groupe[1]
                        
                        # Affichage: Matière, Type, Salle, Groupe
                        type_seance = s[2] if s[2] else "?"
                        titre = s[1] if s[1] else "?"
                        
                        self.set_course_detailed(
                            self.schedule_table, row, day_idx, 
                            titre, type_seance, salle_name, groupe_name
                        )
                        
        except Exception as e:
            print(f"Schedule load error: {e}")

    def set_course_detailed(self, table, row, col, subject, seance_type, room, group):
        """Affiche une séance avec tous les détails: matière, type, salle, groupe"""
        # Couleur selon le type
        color_map = {
            'Cours': COLORS['primary_blue'],
            'TD': '#27ae60',  # Vert
            'TP': '#e67e22',  # Orange
            'Examen': '#c0392b',  # Rouge
        }
        color = color_map.get(seance_type, COLORS['secondary_blue'])
        
        # Widget personnalisé avec tous les détails
        content = f"""
        <div style='text-align: center; padding: 4px;'>
            <b>{subject}</b><br/>
            <span style='font-size: 11px; color: #ffffff; background-color: rgba(255,255,255,0.2); padding: 2px 5px; border-radius: 3px;'>{seance_type}</span><br/>
            <span style='font-size: 10px;'>📍 {room}</span><br/>
            <span style='font-size: 10px;'>👥 {group if group else 'N/A'}</span>
        </div>
        """
        
        item = QLabel(content)
        item.setTextFormat(Qt.TextFormat.RichText)
        item.setAlignment(Qt.AlignmentFlag.AlignCenter)
        item.setStyleSheet(f"""
            background-color: {color}; 
            color: white; 
            border-radius: 8px; 
            margin: 2px; 
            padding: 5px;
        """)
        table.setCellWidget(row, col, item)

    def export_schedule(self, format_type):
        """Exporte l'emploi du temps dans le format demandé"""
        try:
            # Créer le dossier exports s'il n'existe pas
            exports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'exports')
            os.makedirs(exports_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"emploi_du_temps_{self.user.nom}_{timestamp}"
            
            if format_type.lower() == "pdf":
                success, filepath = self._export_to_pdf(exports_dir, filename)
            elif format_type.lower() == "excel":
                success, filepath = self._export_to_excel(exports_dir, filename)
            elif format_type.lower() == "image":
                success, filepath = self._export_to_image(exports_dir, filename)
            else:
                success, filepath = False, None
            
            if success:
                QMessageBox.information(
                    self, 
                    "Export Réussi", 
                    f"✅ Emploi du temps exporté avec succès!\n\nFichier: {filepath}"
                )
            else:
                QMessageBox.warning(self, "Export", f"L'export {format_type} nécessite des dépendances supplémentaires.")
                
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"❌ Erreur lors de l'export: {e}")

    def _export_to_pdf(self, exports_dir, filename):
        """Export vers PDF en utilisant reportlab si disponible, sinon texte"""
        filepath = os.path.join(exports_dir, f"{filename}.pdf")
        
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.pdfgen import canvas
            from reportlab.lib.colors import HexColor
            
            c = canvas.Canvas(filepath, pagesize=landscape(A4))
            width, height = landscape(A4)
            
            # Titre
            c.setFont("Helvetica-Bold", 20)
            c.drawCentredString(width/2, height - 50, f"Emploi du Temps - {self.user.prenom} {self.user.nom}")
            c.setFont("Helvetica", 12)
            c.drawCentredString(width/2, height - 70, f"Spécialité: {getattr(self.user, 'specialite', 'N/A')}")
            
            # Table headers
            days = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi"]
            time_slots = ["08:30-10:00", "10:15-11:45", "12:00-13:30", "13:45-15:15", "15:30-17:00"]
            
            start_y = height - 100
            cell_width = (width - 150) / 6
            cell_height = 80
            
            # Draw headers
            c.setFont("Helvetica-Bold", 10)
            for i, day in enumerate(days):
                x = 100 + i * cell_width
                c.drawString(x + 20, start_y, day)
            
            # Draw time slots
            c.setFont("Helvetica", 9)
            for i, slot in enumerate(time_slots):
                y = start_y - 30 - (i * cell_height)
                c.drawString(10, y, slot)
            
            c.save()
            return True, filepath
            
        except ImportError:
            # Fallback to text file
            text_path = os.path.join(exports_dir, f"{filename}.txt")
            with open(text_path, 'w', encoding='utf-8') as f:
                f.write(f"EMPLOI DU TEMPS - {self.user.prenom} {self.user.nom}\n")
                f.write(f"Spécialité: {getattr(self.user, 'specialite', 'N/A')}\n")
                f.write("="*50 + "\n\n")
                
                seances = self.db.get_seances_by_enseignant(self.user.id)
                for s in seances:
                    salle = self.db.get_salle_by_id(s[6]) if s[6] else None
                    groupe = self.db.get_groupe_by_id(s[8]) if s[8] else None
                    f.write(f"Date: {s[3]} | {s[4]}-{s[5]}\n")
                    f.write(f"  Matière: {s[1]} ({s[2]})\n")
                    f.write(f"  Salle: {salle[1] if salle else 'N/A'}\n")
                    f.write(f"  Groupe: {groupe[1] if groupe else 'N/A'}\n\n")
                    
            return True, text_path

    def _export_to_excel(self, exports_dir, filename):
        """Export vers Excel en utilisant openpyxl si disponible, sinon CSV"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Fill, PatternFill, Alignment
            
            filepath = os.path.join(exports_dir, f"{filename}.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.title = "Emploi du Temps"
            
            # Titre
            ws['A1'] = f"Emploi du Temps - {self.user.prenom} {self.user.nom}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"Spécialité: {getattr(self.user, 'specialite', 'N/A')}"
            
            # Headers
            headers = ["Date", "Horaire", "Matière", "Type", "Salle", "Groupe"]
            for i, header in enumerate(headers):
                cell = ws.cell(row=4, column=i+1, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1e3a8a")
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            seances = self.db.get_seances_by_enseignant(self.user.id)
            for row_num, s in enumerate(seances, start=5):
                salle = self.db.get_salle_by_id(s[6]) if s[6] else None
                groupe = self.db.get_groupe_by_id(s[8]) if s[8] else None
                ws.cell(row=row_num, column=1, value=s[3])
                ws.cell(row=row_num, column=2, value=f"{s[4]}-{s[5]}")
                ws.cell(row=row_num, column=3, value=s[1])
                ws.cell(row=row_num, column=4, value=s[2])
                ws.cell(row=row_num, column=5, value=salle[1] if salle else 'N/A')
                ws.cell(row=row_num, column=6, value=groupe[1] if groupe else 'N/A')
            
            wb.save(filepath)
            return True, filepath
            
        except ImportError:
            # Fallback to CSV
            filepath = os.path.join(exports_dir, f"{filename}.csv")
            
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([f"Emploi du Temps - {self.user.prenom} {self.user.nom}"])
                writer.writerow([f"Spécialité: {getattr(self.user, 'specialite', 'N/A')}"])
                writer.writerow([])
                writer.writerow(["Date", "Horaire", "Matière", "Type", "Salle", "Groupe"])
                
                seances = self.db.get_seances_by_enseignant(self.user.id)
                for s in seances:
                    salle = self.db.get_salle_by_id(s[6]) if s[6] else None
                    groupe = self.db.get_groupe_by_id(s[8]) if s[8] else None
                    writer.writerow([
                        s[3], f"{s[4]}-{s[5]}", s[1], s[2],
                        salle[1] if salle else 'N/A',
                        groupe[1] if groupe else 'N/A'
                    ])
                    
            return True, filepath

    def _export_to_image(self, exports_dir, filename):
        """Export vers Image en capturant le widget de la table"""
        filepath = os.path.join(exports_dir, f"{filename}.png")
        
        try:
            # Capturer le widget de la table
            pixmap = self.schedule_table.grab()
            pixmap.save(filepath, "PNG")
            return True, filepath
        except Exception as e:
            print(f"Image export error: {e}")
            return False, None

    def create_reservation_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        
        # Info enseignant avec spécialité
        info_frame = QFrame()
        info_frame.setStyleSheet("""
            background-color: white; 
            border-radius: 10px; 
            padding: 15px;
            border: 1px solid #E0E0E0;
        """)
        info_layout = QHBoxLayout(info_frame)
        
        specialite = getattr(self.user, 'specialite', None) or "Non définie"
        info_str = f"""
        <div style='font-size: 14px; color: {COLORS['text_dark']};'>
            <b>Enseignant:</b> {self.user.prenom} {self.user.nom}<br/>
            <b>Email:</b> {self.user.email}<br/>
            <b>Spécialité:</b> {specialite}
        </div>
        """
        info_label = QLabel(info_str)
        info_label.setTextFormat(Qt.TextFormat.RichText)
        info_layout.addWidget(info_label)
        info_layout.addStretch()
        layout.addWidget(info_frame)
        
        # Formulaire de réservation
        opt2_frame = QFrame()
        opt2_frame.setStyleSheet(CARD_STYLE)
        opt2_layout = QVBoxLayout(opt2_frame)
        
        title_lbl = QLabel("Nouvelle Demande de Réservation")
        title_lbl.setStyleSheet(CARD_TITLE_STYLE)
        opt2_layout.addWidget(title_lbl)
        
        grid = QGridLayout()
        grid.setSpacing(15)
        
        # Date
        grid.addWidget(QLabel("Date:"), 0, 0)
        self.new_res_date = QDateEdit(QDate.currentDate().addDays(1))
        self.new_res_date.setCalendarPopup(True)
        self.new_res_date.setStyleSheet(INPUT_STYLE)
        self.new_res_date.setMinimumDate(QDate.currentDate())
        grid.addWidget(self.new_res_date, 0, 1)
        
        # Heure début
        grid.addWidget(QLabel("Heure Début:"), 0, 2)
        self.new_res_time_start = QTimeEdit(QTime(8, 30))
        self.new_res_time_start.setStyleSheet(INPUT_STYLE)
        grid.addWidget(self.new_res_time_start, 0, 3)
        
        # Heure fin
        grid.addWidget(QLabel("Heure Fin:"), 0, 4)
        self.new_res_time_end = QTimeEdit(QTime(10, 0))
        self.new_res_time_end.setStyleSheet(INPUT_STYLE)
        grid.addWidget(self.new_res_time_end, 0, 5)
        
        # Matière
        grid.addWidget(QLabel("Matière:"), 1, 0)
        self.new_res_subject = QComboBox()
        self.new_res_subject.setEditable(True)
        self.load_matieres_for_specialite()
        self.new_res_subject.setStyleSheet(INPUT_STYLE)
        grid.addWidget(self.new_res_subject, 1, 1)
        
        # Type
        grid.addWidget(QLabel("Type:"), 1, 2)
        self.new_res_type = QComboBox()
        self.new_res_type.addItems(["Cours", "TD", "TP", "Examen", "Rattrapage"])
        self.new_res_type.setStyleSheet(INPUT_STYLE)
        grid.addWidget(self.new_res_type, 1, 3)
        
        # Salle
        grid.addWidget(QLabel("Salle:"), 1, 4)
        self.new_res_salle = QComboBox()
        self.load_salles()
        self.new_res_salle.setStyleSheet(INPUT_STYLE)
        grid.addWidget(self.new_res_salle, 1, 5)
        
        # Motif
        grid.addWidget(QLabel("Motif:"), 2, 0)
        self.new_res_motif = QTextEdit()
        self.new_res_motif.setPlaceholderText("Raison de la réservation...")
        self.new_res_motif.setStyleSheet(INPUT_STYLE)
        self.new_res_motif.setMaximumHeight(60)
        grid.addWidget(self.new_res_motif, 2, 1, 1, 5)
        
        opt2_layout.addLayout(grid)
        
        # Bouton de soumission
        btn_layout = QHBoxLayout()
        btn_reserve = QPushButton("Soumettre la Demande")
        btn_reserve.setStyleSheet(PRIMARY_BUTTON_STYLE)
        btn_reserve.clicked.connect(self.submit_reservation)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_reserve)
        opt2_layout.addLayout(btn_layout)
        
        layout.addWidget(opt2_frame)
        layout.addStretch()
        return page

    def load_matieres_for_specialite(self):
        """Charge les matières correspondant à la spécialité de l'enseignant"""
        self.new_res_subject.clear()
        specialite = getattr(self.user, 'specialite', None)
        
        matieres_set = set()
        
        if specialite:
            specialite_lower = specialite.lower()
            # Chercher les mots-clés correspondants
            keywords = []
            for spec, kw_list in SPECIALITE_KEYWORDS.items():
                if specialite_lower in spec.lower() or spec.lower() in specialite_lower:
                    keywords.extend(kw_list)
            
            # Chercher les matières correspondantes
            for prog, matieres in MATIERES_COMPLETES.items():
                for mat in matieres:
                    mat_name = mat[1] if isinstance(mat, tuple) else mat
                    mat_lower = mat_name.lower()
                    # Vérifier si la matière correspond
                    for kw in keywords:
                        if kw.lower() in mat_lower:
                            matieres_set.add(mat_name)
                            break
        
        # Si pas de correspondance, ajouter quelques matières génériques
        if not matieres_set:
            matieres_set = {"Cours Magistral", "Travaux Dirigés", "Travaux Pratiques"}
        
        self.new_res_subject.addItems(sorted(matieres_set))

    def load_salles(self):
        """Charge la liste des salles disponibles"""
        self.new_res_salle.clear()
        try:
            salles = self.db.get_toutes_salles()
            for s in salles:
                # s: (id, nom, capacite, type_salle, equipements)
                self.new_res_salle.addItem(f"{s[1]} ({s[3]}, {s[2]} places)", s[0])
        except Exception as e:
            print(f"Erreur chargement salles: {e}")
            self.new_res_salle.addItem("Aucune salle disponible", -1)

    def submit_reservation(self):
        """Soumet une demande de réservation avec validation de spécialité"""
        try:
            # Récupérer les données
            date = self.new_res_date.date().toString("yyyy-MM-dd")
            heure_debut = self.new_res_time_start.time().toString("HH:mm")
            heure_fin = self.new_res_time_end.time().toString("HH:mm")
            matiere = self.new_res_subject.currentText().strip()
            type_seance = self.new_res_type.currentText()
            salle_id = self.new_res_salle.currentData()
            motif = self.new_res_motif.toPlainText()
            
            # Validation de la matière
            if not matiere:
                QMessageBox.warning(self, "Erreur", "Veuillez saisir une matière.")
                return
            
            # Validation de l'heure
            if self.new_res_time_start.time() >= self.new_res_time_end.time():
                QMessageBox.warning(self, "Erreur", "L'heure de fin doit être après l'heure de début.")
                return
            
            # Validation de la salle
            if salle_id is None or salle_id == -1:
                QMessageBox.warning(self, "Erreur", "Veuillez sélectionner une salle valide.")
                return
            
            # Validation de la spécialité
            is_valid, msg = self.db.valider_specialite_enseignant(self.user.id, matiere)
            if not is_valid:
                QMessageBox.warning(
                    self, 
                    "Spécialité Non Correspondante", 
                    f"⚠️ {msg}\n\nVeuillez choisir une matière correspondant à votre spécialité."
                )
                return
            
            # Construire le motif complet
            full_motif = f"{type_seance} - {matiere}"
            if motif:
                full_motif += f" | {motif}"
            
            # Créer la réservation
            res_id = self.db.ajouter_reservation(
                enseignant_id=self.user.id,
                salle_id=salle_id,
                date=date,
                heure_debut=heure_debut,
                heure_fin=heure_fin,
                motif=full_motif
            )
            
            if res_id:
                QMessageBox.information(
                    self, 
                    "Succès", 
                    f"✅ Demande de réservation envoyée avec succès!\n\n"
                    f"• Date: {date}\n"
                    f"• Horaire: {heure_debut} - {heure_fin}\n"
                    f"• Matière: {matiere}\n"
                    f"• Type: {type_seance}\n\n"
                    f"Votre demande sera traitée par l'administration."
                )
                # Réinitialiser le formulaire
                self.new_res_motif.clear()
            else:
                QMessageBox.critical(self, "Erreur", "❌ Erreur lors de la création de la demande.")
                
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"❌ Erreur lors de la soumission: {e}")

    def create_unavailability_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        
        info = QLabel("Signaler une absence ou indisponibilité")
        info.setStyleSheet(CARD_TITLE_STYLE)
        layout.addWidget(info)
        
        # Formulaire
        form_frame = QFrame()
        form_frame.setStyleSheet(CARD_STYLE)
        form_layout = QGridLayout(form_frame)
        form_layout.setSpacing(20)
        
        # Date Début
        form_layout.addWidget(QLabel("Date Début:"), 0, 0)
        self.ab_date_start = QDateEdit(QDate.currentDate())
        self.ab_date_start.setCalendarPopup(True)
        self.ab_date_start.setStyleSheet(INPUT_STYLE)
        form_layout.addWidget(self.ab_date_start, 0, 1)

        # Date Fin (Optional, for range)
        form_layout.addWidget(QLabel("Date Fin (Optionnel):"), 0, 2)
        self.ab_date_end = QDateEdit(QDate.currentDate())
        self.ab_date_end.setCalendarPopup(True)
        self.ab_date_end.setStyleSheet(INPUT_STYLE)
        form_layout.addWidget(self.ab_date_end, 0, 3)
        
        # Checkbox Journée Entière
        from PyQt6.QtWidgets import QCheckBox
        self.chk_full_day = QCheckBox("Journée Entière")
        self.chk_full_day.setStyleSheet("font-size: 14px;")
        self.chk_full_day.setChecked(True)
        form_layout.addWidget(self.chk_full_day, 1, 0, 1, 2)
        
        # Heure Début / Fin (Hidden if full day)
        self.lbl_h_start = QLabel("Heure Début:")
        self.ab_time_start = QTimeEdit(QTime(8, 30))
        self.ab_time_start.setStyleSheet(INPUT_STYLE)
        
        self.lbl_h_end = QLabel("Heure Fin:")
        self.ab_time_end = QTimeEdit(QTime(18, 30))
        self.ab_time_end.setStyleSheet(INPUT_STYLE)
        
        form_layout.addWidget(self.lbl_h_start, 2, 0)
        form_layout.addWidget(self.ab_time_start, 2, 1)
        form_layout.addWidget(self.lbl_h_end, 2, 2)
        form_layout.addWidget(self.ab_time_end, 2, 3)
        
        # Logic visibility
        def toggle_time():
            visible = not self.chk_full_day.isChecked()
            self.lbl_h_start.setVisible(visible)
            self.ab_time_start.setVisible(visible)
            self.lbl_h_end.setVisible(visible)
            self.ab_time_end.setVisible(visible)
            
        self.chk_full_day.toggled.connect(toggle_time)
        toggle_time() # Init
        
        form_layout.addWidget(QLabel("Motif:"), 3, 0)
        self.ab_reason = QTextEdit()
        self.ab_reason.setStyleSheet(INPUT_STYLE)
        self.ab_reason.setMaximumHeight(80)
        form_layout.addWidget(self.ab_reason, 3, 1, 1, 3)
        
        btn_submit = QPushButton("Signaler Indisponibilité")
        btn_submit.setStyleSheet(DANGER_BUTTON_STYLE)
        btn_submit.clicked.connect(self.save_unavailability)
        form_layout.addWidget(btn_submit, 4, 1, 1, 2)
        
        layout.addWidget(form_frame)
        layout.addStretch()
        return page

    def save_unavailability(self):
        d_start = self.ab_date_start.date().toString("yyyy-MM-dd")
        d_end = self.ab_date_end.date().toString("yyyy-MM-dd") 
        reason = self.ab_reason.toPlainText()
        
        # Validation Date
        if self.ab_date_start.date() > self.ab_date_end.date():
             QMessageBox.warning(self, "Erreur", "La date de fin doit être postérieure à la date de début.")
             return

        # Construction du motif avec les périodes
        full_msg = reason
        period_msg = ""
        
        # Période de dates
        if d_start != d_end:
            period_msg = f"[Période: {d_start} -> {d_end}]"
        else:
            period_msg = f"[Date: {d_start}]"

        # Période d'heures
        if not self.chk_full_day.isChecked():
            t_start = self.ab_time_start.time().toString("HH:mm")
            t_end = self.ab_time_end.time().toString("HH:mm")
            
            if self.ab_time_start.time() >= self.ab_time_end.time():
                QMessageBox.warning(self, "Erreur", "L'heure de fin doit être postérieure à l'heure de début.")
                return
                
            period_msg += f" [Heure: {t_start} -> {t_end}]"
        else:
            period_msg += " [Journée entière]"
            
        final_motif = f"{period_msg} {full_msg}"
            
        try:
             conn = self.db.get_connection()
             cursor = conn.cursor()
             
             # Tentative d'insertion avec gestion adaptative des colonnes
             try:
                 # Le schéma standard a date_debut ET date_fin (NOT NULL)
                 cursor.execute('''
                    INSERT INTO disponibilites (enseignant_id, date_debut, date_fin, motif) 
                    VALUES (?, ?, ?, ?)
                 ''', (self.user.id, d_start, d_end, final_motif))
             except Exception as e:
                 # Fallback sur 'date' si la colonne 'date_debut' n'existe pas (ancienne version)
                 # Note: Si date_fin manque, cela echouera aussi si on essaie d'inserer date_fin.
                 # On suppose ici un fallback simple.
                 cursor.execute('''
                    INSERT INTO disponibilites (enseignant_id, date, motif) 
                    VALUES (?, ?, ?)
                 ''', (self.user.id, d_start, final_motif))
                 
             conn.commit()
             conn.close()
             QMessageBox.information(self, "Succès", "Votre indisponibilité a été enregistrée avec succès.")
             self.ab_reason.clear()
        except Exception as e:
             QMessageBox.warning(self, "Erreur", f"Erreur lors de l'enregistrement: {e}")

    def switch_page(self, page_id):
        titles = {
            "Schedule": "Mon Emploi du Temps",
            "Reservation": "Réserver une Salle",
            # "Search": "Rechercher une Salle Vide", # Removed
            "Unavailability": "Mes Indisponibilités"
        }
        self.page_title.setText(titles.get(page_id, page_id))
        
        # Correct mapping based on create_content_area order:
        # 0: Schedule, 1: Reservation, 2: Unavailability
        mapping = {"Schedule": 0, "Reservation": 1, "Unavailability": 2}
        
        if page_id in mapping:
            self.pages.setCurrentIndex(mapping[page_id])
            
        for pid, btn in self.menu_buttons.items():
            btn.setChecked(pid == page_id)

import sys
import os
